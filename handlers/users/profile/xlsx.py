import os

import openpyxl
from aiogram.types import CallbackQuery
from openpyxl.styles import Font, Alignment, Side, Border

from keyboards.inline.callback_datas import set_command_seller_id, choice_date_xlsx_callback
from keyboards.inline.profile_keyboard.seller_settings import back_to_seller_setting, choice_date_xlsx
from loader import dp
from utils.db_api.quick_commands.product_inquiries import get_days_stats
from utils.db_api.quick_commands.seller_inquiries import select_seller


@dp.callback_query_handler(set_command_seller_id.filter(command_name="exel_statistics"))
async def xlsx_stats(call: CallbackQuery, callback_data: dict):
    seller_id = int(callback_data.get("seller_id"))
    await call.message.edit_text("Выберете период за который хотите получить статистику",
                                 reply_markup=choice_date_xlsx(seller_id))


@dp.callback_query_handler(choice_date_xlsx_callback.filter(command_name="show_stats_x"))
async def xlsx_stats(call: CallbackQuery, callback_data: dict):
    await call.message.delete()
    seller_id = int(callback_data.get("seller_id"))
    days = int(callback_data.get("days"))
    seller = await select_seller(seller_id=int(seller_id))
    if seller.tarif is False:
        return await call.message.edit_text("\n".join(
            [
                f'❌ Оплатите подписку для поставщика {seller.name},'
                f'для того чтобы воспользоваться этой функцией',
            ]
        ), reply_markup=back_to_seller_setting(seller_id=seller.id))
    file_name = f"{seller.name}"
    orders, sales_refunds, stocks = await get_days_stats(seller_id, days)
    sales = [sale for sale in sales_refunds if sale.saleID.startswith('S')]
    refunds = [sale for sale in sales_refunds if sale.saleID.startswith('R')]
    await generate_xlsx(sales, orders, refunds, stocks, file_name)
    file_path = f"xlsx/{file_name}.xlsx"
    file = open(file_path, "rb")
    await call.message.bot.send_document(
        chat_id=call.message.chat.id,
        document=file,
        caption="\n".join(
            [
                f'Все данные находятся в файле 👆'
            ]
        )
    )
    if os.path.isfile(file_path):
        os.remove(file_path)


async def generate_xlsx(sales, orders, refunds, stocks, file_name):
    book = openpyxl.Workbook()
    for sheet_name in book.sheetnames:
        sheet = book.get_sheet_by_name(sheet_name)
        book.remove_sheet(sheet)
    order_sheet = book.create_sheet('Заказы')
    sale_sheet = book.create_sheet('Продажи(выкупы)')
    refunds_sheet = book.create_sheet('Возвраты')
    stocks_sheet = book.create_sheet('Остатки')
    font_title = Font(
        name='Arial',
        size=13,
        bold=True,
    )
    font_all = Font(
        name='Arial',
        size=13,
    )
    text_style = Alignment(horizontal='center',
                           vertical='top',
                           text_rotation=0,
                           wrap_text=True,
                           shrink_to_fit=False,
                           indent=0)
    thins = Side(border_style="thin")
    order_sheet.cell(row=1, column=1).value = "Артикул"
    order_sheet.cell(row=1, column=2).value = "Бренд"
    order_sheet.cell(row=1, column=3).value = "Категория"
    order_sheet.cell(row=1, column=4).value = "Подкатегория"
    order_sheet.cell(row=1, column=5).value = "Дата совершения заказа"
    order_sheet.cell(row=1, column=6).value = "Область"
    order_sheet.cell(row=1, column=7).value = "Цена"
    for number in range(1, 8):
        order_sheet.cell(row=1, column=number).font = font_title
        order_sheet.cell(row=1, column=number).border = Border(top=thins, bottom=thins, left=thins, right=thins)
    order_sheet.column_dimensions['A'].width = 36.5
    order_sheet.column_dimensions['B'].width = 36.5
    order_sheet.column_dimensions['C'].width = 36.5
    order_sheet.column_dimensions['D'].width = 36.5
    order_sheet.column_dimensions['E'].width = 36.5
    order_sheet.column_dimensions['F'].width = 36.5
    order_sheet.column_dimensions['G'].width = 36.5
    order_row = 2
    for order in orders:
        order_sheet[order_row][0].value = order.nmId
        order_sheet[order_row][1].value = order.brand
        order_sheet[order_row][2].value = order.category
        order_sheet[order_row][3].value = order.subject
        order_sheet[order_row][4].value = order.date.strftime("%d.%m.%Y %H:%M")
        order_sheet[order_row][5].value = order.oblast
        order_sheet[order_row][6].value = order.price
        # шрифт
        order_sheet[order_row][0].font = font_all
        order_sheet[order_row][1].font = font_all
        order_sheet[order_row][2].font = font_all
        order_sheet[order_row][3].font = font_all
        order_sheet[order_row][4].font = font_all
        order_sheet[order_row][5].font = font_all
        order_sheet[order_row][6].font = font_all
        # Выравнивание текста
        order_sheet[order_row][0].alignment = text_style
        order_sheet[order_row][1].alignment = text_style
        order_sheet[order_row][2].alignment = text_style
        order_sheet[order_row][3].alignment = text_style
        order_sheet[order_row][4].alignment = text_style
        order_sheet[order_row][5].alignment = text_style
        order_row += 1

    sale_sheet.cell(row=1, column=1).value = "Артикул"
    sale_sheet.cell(row=1, column=2).value = "Бренд"
    sale_sheet.cell(row=1, column=3).value = "Категория"
    sale_sheet.cell(row=1, column=4).value = "Подкатегория"
    sale_sheet.cell(row=1, column=5).value = "Дата совершения заказа"
    sale_sheet.cell(row=1, column=6).value = "Регион"
    sale_sheet.cell(row=1, column=7).value = "Цена"
    for number in range(1, 8):
        sale_sheet.cell(row=1, column=number).font = font_title
        sale_sheet.cell(row=1, column=number).border = Border(top=thins, bottom=thins, left=thins, right=thins)
    sale_sheet.column_dimensions['A'].width = 36.5
    sale_sheet.column_dimensions['B'].width = 36.5
    sale_sheet.column_dimensions['C'].width = 36.5
    sale_sheet.column_dimensions['D'].width = 36.5
    sale_sheet.column_dimensions['E'].width = 36.5
    sale_sheet.column_dimensions['F'].width = 36.5
    sale_sheet.column_dimensions['G'].width = 36.5
    sale_row = 2
    for sale in sales:
        sale_sheet[sale_row][0].value = sale.nmId
        sale_sheet[sale_row][1].value = sale.brand
        sale_sheet[sale_row][2].value = sale.category
        sale_sheet[sale_row][3].value = sale.subject
        sale_sheet[sale_row][4].value = sale.date.strftime("%d.%m.%Y %H:%M")
        sale_sheet[sale_row][5].value = sale.regionName
        sale_sheet[sale_row][6].value = sale.forPay
        # шрифт
        sale_sheet[sale_row][0].font = font_all
        sale_sheet[sale_row][1].font = font_all
        sale_sheet[sale_row][2].font = font_all
        sale_sheet[sale_row][3].font = font_all
        sale_sheet[sale_row][4].font = font_all
        sale_sheet[sale_row][5].font = font_all
        sale_sheet[sale_row][6].font = font_all
        # Выравнивание текста
        sale_sheet[sale_row][0].alignment = text_style
        sale_sheet[sale_row][1].alignment = text_style
        sale_sheet[sale_row][2].alignment = text_style
        sale_sheet[sale_row][3].alignment = text_style
        sale_sheet[sale_row][4].alignment = text_style
        sale_sheet[sale_row][5].alignment = text_style
        sale_row += 1

    refunds_sheet.cell(row=1, column=1).value = "Артикул"
    refunds_sheet.cell(row=1, column=2).value = "Бренд"
    refunds_sheet.cell(row=1, column=3).value = "Категория"
    refunds_sheet.cell(row=1, column=4).value = "Подкатегория"
    refunds_sheet.cell(row=1, column=5).value = "Дата совершения заказа"
    refunds_sheet.cell(row=1, column=6).value = "Регион"
    refunds_sheet.cell(row=1, column=7).value = "Сумма возврата"
    for number in range(1, 8):
        refunds_sheet.cell(row=1, column=number).font = font_title
        refunds_sheet.cell(row=1, column=number).border = Border(top=thins, bottom=thins, left=thins, right=thins)
    refunds_sheet.column_dimensions['A'].width = 36.5
    refunds_sheet.column_dimensions['B'].width = 36.5
    refunds_sheet.column_dimensions['C'].width = 36.5
    refunds_sheet.column_dimensions['D'].width = 36.5
    refunds_sheet.column_dimensions['E'].width = 36.5
    refunds_sheet.column_dimensions['F'].width = 36.5
    refunds_sheet.column_dimensions['G'].width = 36.5
    refund_row = 2
    for refund in refunds:
        refunds_sheet[refund_row][0].value = refund.nmId
        refunds_sheet[refund_row][1].value = refund.brand
        refunds_sheet[refund_row][2].value = refund.category
        refunds_sheet[refund_row][3].value = refund.subject
        refunds_sheet[refund_row][4].value = refund.date.strftime("%d.%m.%Y %H:%M")
        refunds_sheet[refund_row][5].value = refund.regionName
        refunds_sheet[refund_row][6].value = refund.forPay
        # шрифт
        refunds_sheet[refund_row][0].font = font_all
        refunds_sheet[refund_row][1].font = font_all
        refunds_sheet[refund_row][2].font = font_all
        refunds_sheet[refund_row][3].font = font_all
        refunds_sheet[refund_row][4].font = font_all
        refunds_sheet[refund_row][5].font = font_all
        refunds_sheet[refund_row][6].font = font_all
        # Выравнивание текста
        refunds_sheet[refund_row][0].alignment = text_style
        refunds_sheet[refund_row][1].alignment = text_style
        refunds_sheet[refund_row][2].alignment = text_style
        refunds_sheet[refund_row][3].alignment = text_style
        refunds_sheet[refund_row][4].alignment = text_style
        refunds_sheet[refund_row][5].alignment = text_style
        refund_row += 1

    stocks_sheet.cell(row=1, column=1).value = "Артикул"
    stocks_sheet.cell(row=1, column=2).value = "Бренд"
    stocks_sheet.cell(row=1, column=3).value = "Категория"
    stocks_sheet.cell(row=1, column=4).value = "Подкатегория"
    stocks_sheet.cell(row=1, column=5).value = "В наличии"
    stocks_sheet.cell(row=1, column=6).value = "В пути до клиента"
    stocks_sheet.cell(row=1, column=7).value = "В пути от клиента"
    stocks_sheet.cell(row=1, column=8).value = "Склад"
    for number in range(1, 9):
        stocks_sheet.cell(row=1, column=number).font = font_title
        stocks_sheet.cell(row=1, column=number).border = Border(top=thins, bottom=thins, left=thins, right=thins)
    stocks_sheet.column_dimensions['A'].width = 36.5
    stocks_sheet.column_dimensions['B'].width = 36.5
    stocks_sheet.column_dimensions['C'].width = 36.5
    stocks_sheet.column_dimensions['D'].width = 36.5
    stocks_sheet.column_dimensions['E'].width = 36.5
    stocks_sheet.column_dimensions['F'].width = 36.5
    stocks_sheet.column_dimensions['G'].width = 36.5
    stocks_sheet.column_dimensions['H'].width = 36.5
    stock_row = 2
    for stock in stocks:
        stocks_sheet[stock_row][0].value = stock.nmId
        stocks_sheet[stock_row][1].value = stock.brand
        stocks_sheet[stock_row][2].value = stock.category
        stocks_sheet[stock_row][3].value = stock.subject
        stocks_sheet[stock_row][4].value = stock.quantityFull
        stocks_sheet[stock_row][5].value = stock.inWayToClient
        stocks_sheet[stock_row][6].value = stock.inWayFromClient
        stocks_sheet[stock_row][7].value = stock.warehouseName
        # шрифт
        stocks_sheet[stock_row][0].font = font_all
        stocks_sheet[stock_row][1].font = font_all
        stocks_sheet[stock_row][2].font = font_all
        stocks_sheet[stock_row][3].font = font_all
        stocks_sheet[stock_row][4].font = font_all
        stocks_sheet[stock_row][5].font = font_all
        stocks_sheet[stock_row][6].font = font_all
        stocks_sheet[stock_row][7].font = font_all
        # Выравнивание текста
        stocks_sheet[stock_row][0].alignment = text_style
        stocks_sheet[stock_row][1].alignment = text_style
        stocks_sheet[stock_row][2].alignment = text_style
        stocks_sheet[stock_row][3].alignment = text_style
        stocks_sheet[stock_row][4].alignment = text_style
        stocks_sheet[stock_row][5].alignment = text_style
        stocks_sheet[stock_row][6].alignment = text_style
        stocks_sheet[stock_row][7].alignment = text_style
        stock_row += 1
    book.save(f"xlsx/{file_name}.xlsx")
    book.close()
